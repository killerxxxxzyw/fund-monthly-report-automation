# -*- coding: utf-8 -*-
"""Online spreadsheet writer with dry-run safety checks.

This is a sanitized public version of the browser automation layer. It retains
fixed-cell navigation, read-before-write comparison, vertical block pasting,
retry logic and structural-risk blocking, while excluding production document
URLs, login state and proprietary portfolio names.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import openpyxl

try:
    import pyperclip
except ModuleNotFoundError:
    pyperclip = None

try:
    from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright
except ModuleNotFoundError:
    Page = Any
    PlaywrightTimeoutError = TimeoutError
    sync_playwright = None

from config import (
    CHROME_PATH, DOC_URL, EXPECTED_MARKET_VALUE_ROWS, LOGIN_STATE_DIR,
    MARKET_VALUE_START_CELLS, OUTPUT_FILE, PORTFOLIO_ORDER, SUMMARY_CELL_MAP,
    TARGET_SHEET_KEYWORD,
)

BASE_DIR = Path(__file__).resolve().parent
LOG_FILE = BASE_DIR / "output" / "playwright_change_plan.txt"
PRODUCT_ORDER = PORTFOLIO_ORDER

NAME_BOX_X = 95
NAME_BOX_Y = 106
DRY_RUN = True
STRUCTURAL_CHANGES_MODE = "manual"

HEADER_CELLS = {
    "净值更新日期": "B1",
    "持仓更新日期": "B2",
    "剩余资金更新": "B3",
    "市值": ["D5", "D15", "D25"],
    "投资进度": ["E5", "E15", "E25"],
    "认购申购情况": ["F5", "F15", "F25"],
    "提取情况": ["G5", "G15", "G25"],
    "已投金额": ["I5", "I15", "I25"],
    "现金分红金额": ["J5", "J15", "J25"],
}

KDOCS_SEQUENCE_CONFIRMED = {portfolio: True for portfolio in PRODUCT_ORDER}

@dataclass
class ReportDates:
    slash: str
    compact: str
    short: str


@dataclass
class ChangeItem:
    product: str
    target_range: str
    row_count: int
    fields: list[str]
    data: list[list[Any]] | list[Any] | Any
    status: str
    safe_to_write: bool
    reason: str


@dataclass
class ProductWriteStats:
    checked_cells: int = 0
    skipped_cells: int = 0
    written_cells: int = 0
    write_operations: int = 0
    failed: int = 0
    skipped_items: list[str] = field(default_factory=list)
    written_items: list[str] = field(default_factory=list)
    failed_items: list[str] = field(default_factory=list)
    diff_items: list[str] = field(default_factory=list)


@dataclass
class WriteStats:
    checked_cells: int = 0
    skipped_cells: int = 0
    written_cells: int = 0
    write_operations: int = 0
    failed: int = 0
    products: dict[str, ProductWriteStats] = field(default_factory=dict)

    def product(self, product: str) -> ProductWriteStats:
        if product not in self.products:
            self.products[product] = ProductWriteStats()
        return self.products[product]


@dataclass
class RangeReadResult:
    values: list[str]
    success: bool


@dataclass
class StructuralChange:
    product: str
    fund_name: str
    current_index: Any
    action: str
    suggested_position: str
    before_count: int
    after_count: int
    status: str


class ValueComparator:
    """比较在线表格显示值与准备写入值，避免重复写入。"""

    FLOAT_TOLERANCE = 1e-6

    @classmethod
    def equals(cls, current: Any, expected: Any) -> bool:
        current_text = cls._to_text(current)
        expected_text = cls._to_text(expected)
        if cls._blank(current_text) and cls._blank(expected_text):
            return True

        current_date = cls._parse_date(current_text)
        expected_date = cls._parse_date(expected_text)
        if current_date and expected_date:
            return current_date == expected_date

        current_number = cls._parse_number(current_text)
        expected_number = cls._parse_number(expected_text)
        if current_number is not None and expected_number is not None:
            return abs(current_number - expected_number) <= cls.FLOAT_TOLERANCE

        return cls._normalize_text(current_text) == cls._normalize_text(expected_text)

    @classmethod
    def _to_text(cls, value: Any) -> str:
        return "" if value is None else str(value)

    @classmethod
    def _blank(cls, value: str) -> bool:
        return cls._normalize_text(value) == ""

    @staticmethod
    def _normalize_text(value: str) -> str:
        return re.sub(r"\s+", "", value.replace("\u3000", " ")).strip()

    @staticmethod
    def _parse_number(value: str) -> float | None:
        text = value.strip().replace(",", "")
        if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _parse_date(value: str) -> str | None:
        text = value.strip()
        match = re.fullmatch(r"(20\d{2})/(\d{1,2})/(\d{1,2})", text)
        if not match:
            return None
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


@dataclass
class LoadedReport:
    dates: ReportDates
    summary_rows: dict[str, dict[str, Any]]
    market_rows: dict[str, list[dict[str, Any]]]
    raw_market_rows: list[dict[str, Any]]
    structural_changes: list[StructuralChange] = field(default_factory=list)


class MonthlyReportLoader:
    """读取 output/月报数据.xlsx，只依赖下游已生成的两个金山填表 sheet。"""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path

    def load(self) -> LoadedReport:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"找不到 {self.workbook_path}，请先运行 main.py 生成月报数据。")

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
        self._require_sheets(wb, ["填表用_组合汇总", "在线表格粘贴版"])

        summary_rows = self._read_rows(wb["填表用_组合汇总"])
        market_rows = self._read_rows(wb["在线表格粘贴版"])

        grouped_summary = {self._text(row.get("组合产品")): row for row in summary_rows}
        grouped_market: dict[str, list[dict[str, Any]]] = {product: [] for product in PRODUCT_ORDER}
        for row in market_rows:
            product = self._text(row.get("组合产品"))
            if product in grouped_market:
                grouped_market[product].append(row)

        structural_changes = self._detect_structural_changes(grouped_market, market_rows)

        return LoadedReport(
            dates=self._detect_dates(wb),
            summary_rows=grouped_summary,
            market_rows=grouped_market,
            raw_market_rows=market_rows,
            structural_changes=structural_changes,
        )

    @staticmethod
    def _require_sheets(wb: openpyxl.Workbook, sheet_names: Iterable[str]) -> None:
        missing = [sheet for sheet in sheet_names if sheet not in wb.sheetnames]
        if missing:
            raise ValueError(f"月报数据.xlsx 缺少 Sheet：{missing}；已有 Sheet：{wb.sheetnames}")

    @staticmethod
    def _read_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
        headers = [MonthlyReportLoader._text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        rows: list[dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            row = {
                header: ws.cell(row_idx, col_idx).value
                for col_idx, header in enumerate(headers, start=1)
                if header
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return rows

    @staticmethod
    def _text(value: Any) -> str:
        return "" if value is None else str(value).replace("\u3000", " ").strip()

    def _detect_dates(self, wb: openpyxl.Workbook) -> ReportDates:
        candidates: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, date):
                        candidates.append(value.strftime("%Y%m%d"))
                    elif isinstance(value, str):
                        candidates.extend(re.findall(r"20\d{6}", value))

        for path in (BASE_DIR / "input").glob("*"):
            candidates.extend(re.findall(r"20\d{6}", path.name))

        if not candidates:
            raise ValueError("无法从月报数据或估值文件名中识别 YYYYMMDD 日期。")

        compact = max(candidates)
        year, month, day = int(compact[:4]), int(compact[4:6]), int(compact[6:8])
        return ReportDates(slash=f"{year}/{month}/{day}", compact=compact, short=compact[4:])

    @classmethod
    def _detect_structural_changes(
        cls,
        grouped_market: dict[str, list[dict[str, Any]]],
        market_rows: list[dict[str, Any]],
    ) -> list[StructuralChange]:
        changes: list[StructuralChange] = []
        for row in market_rows:
            product = cls._text(row.get("组合产品"))
            fund_name = cls._text(row.get("子基金"))
            current_index = row.get("序号", "")
            action = cls._text(row.get("建议动作"))
            status = cls._text(row.get("状态"))
            combined = f"{status} {action}"
            if not cls._is_structural_risk(combined):
                continue

            before_count = len(grouped_market.get(product, []))
            after_count = cls._estimate_after_count(before_count, combined)
            changes.append(
                StructuralChange(
                    product=product,
                    fund_name=fund_name,
                    current_index=current_index,
                    action=action or status,
                    suggested_position=cls._suggest_position(product, current_index, combined),
                    before_count=before_count,
                    after_count=after_count,
                    status=status,
                )
            )
        return changes

    @staticmethod
    def _is_structural_risk(text: str) -> bool:
        keywords = ["新增待插行", "新增一行", "疑似全赎", "剪切到文档最后", "在线表格原有但估值表未持有"]
        broad_keywords = ["新增", "全赎", "剪切", "未持有"]
        return any(keyword in text for keyword in keywords + broad_keywords)

    @staticmethod
    def _estimate_after_count(before_count: int, text: str) -> int:
        if "新增" in text:
            return before_count + 1
        if "全赎" in text or "剪切" in text or "未持有" in text:
            return max(before_count - 1, 0)
        return before_count

    @staticmethod
    def _suggest_position(product: str, current_index: Any, text: str) -> str:
        if "新增" in text:
            index_text = current_index if current_index not in (None, "") else "待人工判断"
            return f"{product} 组内按真实顺序插入，参考当前序号：{index_text}"
        if "全赎" in text or "剪切" in text or "未持有" in text:
            return "剪切到文档最后归档区，并重排原组合产品组内序号"
        return "人工确认插入或移出位置"


class HeaderUpdater:
    """生成表头日期更新计划，位置全部来自 HEADER_CELLS。"""

    def build_plan(self, dates: ReportDates) -> list[ChangeItem]:
        values = {
            "净值更新日期": dates.slash,
            "持仓更新日期": dates.slash,
            "剩余资金更新": dates.slash,
            "市值": f"{dates.short}市值",
            "投资进度": f"投资进度（更新到{dates.compact}）",
            "认购申购情况": f"2026年认购申购情况（更新到{dates.compact}）",
            "提取情况": f"2026年提取情况（更新到{dates.compact}）",
            "已投金额": f"已投金额（成立以来包含所有申赎数据，{dates.compact}）",
            "现金分红金额": f"26年现金分红金额（更新到{dates.compact}）",
        }
        items: list[ChangeItem] = []
        for field_name, value in values.items():
            cells = self._as_cells(HEADER_CELLS.get(field_name, ""))
            if not cells:
                cells = [""]
            for cell in cells:
                items.append(
                    ChangeItem(
                        product="表头",
                        target_range=cell or "未配置",
                        row_count=1,
                        fields=[field_name],
                        data=value,
                        status="待写入" if cell else "阻止写入",
                        safe_to_write=bool(cell),
                        reason="固定表头单元格已配置" if cell else f"HEADER_CELLS 缺少 {field_name} 的目标单元格",
                    )
                )
        return items

    @staticmethod
    def _as_cells(value: Any) -> list[str]:
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        if value:
            return [str(value).strip()]
        return []

    def apply(self, browser: "KDocsBrowser", items: list[ChangeItem]) -> None:
        for item in items:
            browser.write_cell(item.target_range, item.data, item)


class SummaryBlockUpdater:
    """生成并写入组合产品汇总块；不通过搜索产品名定位。"""

    FIELD_MAP = {
        "净值": "净值",
        "产品成本规模": "产品成本规模_实收资本",
        "产品市值规模": "产品市值规模_基金资产净值",
        "剩余资金": "剩余资金_万元",
        "剩余资金备注": "剩余资金备注",
    }

    def build_plan(self, report: LoadedReport) -> list[ChangeItem]:
        items: list[ChangeItem] = []
        for product in PRODUCT_ORDER:
            row = report.summary_rows.get(product, {})
            for field_name, excel_field in self.FIELD_MAP.items():
                cell = SUMMARY_CELL_MAP.get(product, {}).get(field_name, "")
                value = self._format_value(field_name, row.get(excel_field, ""))
                items.append(
                    ChangeItem(
                        product=product,
                        target_range=cell or "未配置",
                        row_count=1,
                        fields=[field_name],
                        data=value,
                        status="待写入" if cell else "阻止写入",
                        safe_to_write=bool(cell),
                        reason="汇总块固定单元格已配置" if cell else f"SUMMARY_CELL_MAP 缺少 {product} - {field_name}",
                    )
                )
        return items

    @staticmethod
    def _format_value(field_name: str, value: Any) -> Any:
        if value in (None, ""):
            return ""
        if field_name in {"产品成本规模", "产品市值规模"}:
            return round(float(value) / 10000)
        if field_name == "剩余资金":
            return round(float(value))
        return value

    def apply(self, browser: "KDocsBrowser", items: list[ChangeItem]) -> None:
        for item in items:
            browser.write_cell(item.target_range, item.data, item)


class MarketValueBlockUpdater:
    """按组合产品固定顺序整列粘贴市值。"""

    def build_plan(self, report: LoadedReport) -> list[ChangeItem]:
        items: list[ChangeItem] = []
        blocked_products = {change.product for change in report.structural_changes}
        for product in PRODUCT_ORDER:
            rows = report.market_rows.get(product, [])
            values = [row.get("市值") for row in rows]
            start_cell = MARKET_VALUE_START_CELLS.get(product, "")
            target = self._target_range(start_cell, len(values))
            expected = EXPECTED_MARKET_VALUE_ROWS.get(product)
            has_structure_risk = product in blocked_products
            sequence_confirmed = KDOCS_SEQUENCE_CONFIRMED.get(product, False)
            row_count_ready = expected is not None and len(values) == expected
            safe_to_write = bool(start_cell) and not has_structure_risk and sequence_confirmed and row_count_ready
            if has_structure_risk:
                status = "阻止写入"
                reason = f"{product} 存在未处理的结构变化，STRUCTURAL_CHANGES_MODE={STRUCTURAL_CHANGES_MODE}，需人工先处理"
            elif not sequence_confirmed:
                status = "阻止写入"
                reason = f"KDOCS_SEQUENCE_CONFIRMED 未确认 {product} 的在线表格现有顺序与 Excel 固定顺序一致"
            elif not start_cell:
                status = "阻止写入"
                reason = f"MARKET_VALUE_START_CELLS 缺少 {product} 的起始单元格"
            elif expected is None:
                status = "阻止写入"
                reason = f"EXPECTED_MARKET_VALUE_ROWS 缺少 {product} 人工确认后的金山现有行数"
            elif len(values) != expected:
                status = "阻止写入"
                reason = f"{product} 数据条数 {len(values)} 与预期行数 {expected} 不一致"
            else:
                status = "待整列粘贴"
                reason = "市值起始单元格已配置，结构无风险，数据条数与预期一致，按 Excel 顺序整列粘贴"
            items.append(
                ChangeItem(
                    product=product,
                    target_range=target,
                    row_count=len(values),
                    fields=["市值"],
                    data=values,
                    status=status,
                    safe_to_write=safe_to_write,
                    reason=reason,
                )
            )
        return items

    @staticmethod
    def _target_range(start_cell: str, row_count: int) -> str:
        if not start_cell:
            return "未配置"
        match = re.fullmatch(r"([A-Z]+)(\d+)", start_cell.strip().upper())
        if not match or row_count <= 0:
            return start_cell
        col, start_row_text = match.groups()
        start_row = int(start_row_text)
        return f"{col}{start_row}:{col}{start_row + row_count - 1}"

    def apply(self, browser: "KDocsBrowser", items: list[ChangeItem]) -> None:
        for item in items:
            start_cell = item.target_range.split(":", 1)[0]
            browser.paste_vertical_block(start_cell, item.data, item)


class SafetyValidator:
    """真实写入前的硬性安全校验。"""

    def validate(self, report: LoadedReport, plan: list[ChangeItem]) -> list[str]:
        issues: list[str] = []
        issues.extend(self._validate_config(plan))
        issues.extend(self._validate_product_rows(report))
        issues.extend(self._validate_market_values(report))
        issues.extend(self._validate_structural_risks(report))
        issues.extend(self._validate_duplicate_keys(report))
        return issues

    @staticmethod
    def _validate_config(plan: list[ChangeItem]) -> list[str]:
        return [f"{item.product} {','.join(item.fields)}：{item.reason}" for item in plan if not item.safe_to_write]

    @staticmethod
    def _validate_product_rows(report: LoadedReport) -> list[str]:
        issues: list[str] = []
        for product in PRODUCT_ORDER:
            actual = len(report.market_rows.get(product, []))
            expected = EXPECTED_MARKET_VALUE_ROWS.get(product)
            if expected is None:
                issues.append(f"EXPECTED_MARKET_VALUE_ROWS 缺少 {product} 人工确认后的金山现有行数")
            elif actual != expected:
                issues.append(f"{product} 市值数据条数 {actual} 与配置预期 {expected} 不一致")
        return issues

    @staticmethod
    def _validate_market_values(report: LoadedReport) -> list[str]:
        issues: list[str] = []
        for product, rows in report.market_rows.items():
            blank_rows = [idx for idx, row in enumerate(rows, start=1) if row.get("市值") in (None, "")]
            if blank_rows:
                issues.append(f"{product} 存在空市值，组内序号：{blank_rows}")
        return issues

    @staticmethod
    def _validate_structural_risks(report: LoadedReport) -> list[str]:
        issues: list[str] = []
        for change in report.structural_changes:
            issues.append(
                f"{change.product} - {change.fund_name} 存在结构变化（状态={change.status}，建议动作={change.action}），"
                "该组合产品禁止整列市值粘贴，需人工处理插行/剪切/重排序号"
            )
        for row in report.raw_market_rows:
            product = row.get("组合产品", "")
            fund = row.get("子基金", "")
            paste_flag = str(row.get("是否可直接粘贴市值") or "")
            if paste_flag and paste_flag != "是":
                issues.append(f"{product} - {fund} 是否可直接粘贴市值={paste_flag}，需人工确认")
        return issues

    @staticmethod
    def _validate_duplicate_keys(report: LoadedReport) -> list[str]:
        counter = Counter((row.get("组合产品"), row.get("子基金")) for row in report.raw_market_rows)
        duplicates = [key for key, count in counter.items() if count > 1]
        return [f"存在重复的 组合产品+子基金 键：{duplicates}"] if duplicates else []


class ComplexFieldPlanner:
    """申购、赎回、分红、已投金额第一版只进入计划，不自动写入。"""

    def build_plan(self, report: LoadedReport) -> list[ChangeItem]:
        grouped = defaultdict(
            lambda: {
                "申购": [],
                "赎回": [],
                "分红": [],
                "已投金额": [],
                "新增": [],
                "疑似全赎": [],
                "剪切到文档最后": [],
                "在线表格原有但估值表未持有": [],
            }
        )
        for row in report.raw_market_rows:
            product = row.get("组合产品", "")
            fund = row.get("子基金", "")
            if (row.get("申购金额") or 0) != 0:
                grouped[product]["申购"].append(fund)
            if (row.get("赎回金额") or 0) != 0:
                grouped[product]["赎回"].append(fund)
            if (row.get("分红金额") or 0) != 0:
                grouped[product]["分红"].append(fund)
            if (row.get("已投金额调整=申购-赎回") or 0) != 0:
                grouped[product]["已投金额"].append(fund)
            action = str(row.get("建议动作") or "")
            status = str(row.get("状态") or "")
            if "新增" in action or "新增" in status:
                grouped[product]["新增"].append(fund)
            if "全赎" in action or "全赎" in status:
                grouped[product]["疑似全赎"].append(fund)
            if "剪切到文档最后" in action or "剪切" in action:
                grouped[product]["剪切到文档最后"].append(fund)
            if "在线表格原有但估值表未持有" in action or "未持有" in status or "未持有" in action:
                grouped[product]["在线表格原有但估值表未持有"].append(fund)

        items: list[ChangeItem] = []
        for product in PRODUCT_ORDER:
            data = grouped[product]
            items.append(
                ChangeItem(
                    product=product,
                    target_range="人工确认",
                    row_count=sum(len(v) for v in data.values()),
                    fields=["申购", "赎回", "分红", "已投金额", "新增", "疑似全赎", "剪切到文档最后", "在线表格原有但估值表未持有"],
                    data=dict(data),
                    status="仅生成计划",
                    safe_to_write=False,
                    reason="复杂字段与结构性操作第一版不自动写入",
                )
            )
        return items


class KDocsBrowser:
    """封装在线表格浏览器操作。只按固定单元格跳转和粘贴，不做全表搜索。"""

    def __init__(self, doc_url: str):
        self.doc_url = doc_url
        self.playwright = None
        self.context = None
        self.page: Page | None = None
        self.stats = WriteStats()

    def __enter__(self) -> "KDocsBrowser":
        if sync_playwright is None:
            raise RuntimeError("缺少 playwright，无法打开在线表格；请先安装依赖后再执行浏览器相关操作。")
        self.playwright = sync_playwright().start()
        launch_kwargs: dict[str, Any] = {
            "user_data_dir": str(LOGIN_STATE_DIR),
            "headless": False,
        }
        if CHROME_PATH and Path(CHROME_PATH).exists():
            launch_kwargs["executable_path"] = str(CHROME_PATH)
        self.context = self.playwright.chromium.launch_persistent_context(**launch_kwargs)
        self.page = self.context.new_page()
        self.page.goto(self.doc_url, wait_until="domcontentloaded")
        self.page.wait_for_timeout(5000)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if self.context:
            self.context.close()
        if self.playwright:
            self.playwright.stop()

    def select_latest_sheet(self) -> bool:
        """尽量点击名称包含目标关键字且日期最新的工作表；失败时交给用户手动切换。"""
        assert self.page is not None
        candidates = self._visible_sheet_candidates()
        if not candidates:
            print(f"未能自动找到包含“{TARGET_SHEET_KEYWORD}”的工作表标签。")
            input("请手动切换到目标工作表后按回车继续：")
            return False

        latest = max(candidates, key=lambda item: item[0])
        try:
            latest[1].click(timeout=3000)
            self.page.wait_for_timeout(800)
            print(f"已自动切换工作表：{latest[2]}")
            return True
        except PlaywrightTimeoutError:
            print(f"自动点击工作表失败：{latest[2]}")
            input("请手动切换到目标工作表后按回车继续：")
            return False

    def _visible_sheet_candidates(self) -> list[tuple[str, Any, str]]:
        assert self.page is not None
        candidates: list[tuple[str, Any, str]] = []
        locator = self.page.locator(f"text={TARGET_SHEET_KEYWORD}")
        for idx in range(locator.count()):
            element = locator.nth(idx)
            text = (element.inner_text(timeout=1000) or "").strip()
            date_key = self._sheet_date_key(text)
            candidates.append((date_key, element, text))
        return candidates

    @staticmethod
    def _sheet_date_key(text: str) -> str:
        compact_dates = re.findall(r"20\d{6}", text)
        if compact_dates:
            return max(compact_dates)
        loose_dates = re.findall(r"20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}", text)
        normalized = []
        for item in loose_dates:
            nums = re.findall(r"\d+", item)
            if len(nums) >= 3:
                normalized.append(f"{int(nums[0]):04d}{int(nums[1]):02d}{int(nums[2]):02d}")
        return max(normalized) if normalized else "00000000"

    def goto_cell(self, cell: str) -> None:
        assert self.page is not None
        self.page.mouse.click(NAME_BOX_X, NAME_BOX_Y)
        self.page.keyboard.press("Control+A")
        self.page.keyboard.type(cell)
        self.page.keyboard.press("Enter")
        self.page.wait_for_timeout(300)

    def write_cell(self, cell: str, value: Any, item: ChangeItem | None = None) -> None:
        product = item.product if item else "未分组"
        field_name = ",".join(item.fields) if item else "单元格"
        product_stats = self.stats.product(product)
        self.stats.checked_cells += 1
        product_stats.checked_cells += 1
        try:
            current = self.read_cell(cell)
            if ValueComparator.equals(current, value):
                self.stats.skipped_cells += 1
                product_stats.skipped_cells += 1
                product_stats.skipped_items.append(f"{cell} {field_name}")
                return

            product_stats.diff_items.append(
                f"{cell} {field_name}：金山当前值={current!r}；新值={self._to_clipboard_text(value)!r}；原因=标准化比较不一致"
            )
            self.goto_cell(cell)
            self.set_clipboard_text(self._to_clipboard_text(value))
            assert self.page is not None
            self.page.keyboard.press("Control+V")
            self.page.wait_for_timeout(200)
            self.stats.written_cells += 1
            self.stats.write_operations += 1
            product_stats.written_cells += 1
            product_stats.write_operations += 1
            product_stats.written_items.append(f"{cell} {field_name}")
        except Exception as exc:
            self.stats.failed += 1
            product_stats.failed += 1
            product_stats.failed_items.append(f"{cell} {field_name}: {exc}")
            raise

    def paste_vertical_block(self, start_cell: str, values: list[Any], item: ChangeItem | None = None) -> None:
        product = item.product if item else "未分组"
        field_name = ",".join(item.fields) if item else "整列"
        product_stats = self.stats.product(product)
        cell_count = len(values)
        self.stats.checked_cells += cell_count
        product_stats.checked_cells += cell_count
        try:
            target_range = item.target_range if item else self._target_range(start_cell, cell_count)
            range_result = self.read_vertical_range(target_range, cell_count)
            if not range_result.success:
                self.stats.failed += 1
                product_stats.failed += 1
                product_stats.failed_items.append(f"{target_range} {field_name}: 范围读取失败，放弃写入")
                return

            current_values = range_result.values
            if len(current_values) == cell_count and all(
                ValueComparator.equals(current, expected)
                for current, expected in zip(current_values, values)
            ):
                self.stats.skipped_cells += cell_count
                product_stats.skipped_cells += cell_count
                product_stats.skipped_items.append(f"{target_range} {field_name}")
                return

            diff_parts = []
            for idx, (current, expected) in enumerate(zip(current_values, values), start=1):
                if not ValueComparator.equals(current, expected):
                    diff_parts.append(
                        f"第{idx}项 金山当前值={current!r} 新值={self._to_clipboard_text(expected)!r}"
                    )
            if len(current_values) != cell_count:
                diff_parts.append(f"读取项数={len(current_values)}，新值项数={cell_count}")
            product_stats.diff_items.append(
                f"{target_range} {field_name}：原因=整列存在不一致；" + "；".join(diff_parts)
            )
            self.goto_cell(start_cell)
            self.set_clipboard_text("\n".join(self._to_clipboard_text(value) for value in values))
            assert self.page is not None
            self.page.keyboard.press("Control+V")
            self.page.wait_for_timeout(500)
            self.stats.written_cells += cell_count
            self.stats.write_operations += 1
            product_stats.written_cells += cell_count
            product_stats.write_operations += 1
            product_stats.written_items.append(f"{target_range} {field_name}")
        except Exception as exc:
            self.stats.failed += 1
            product_stats.failed += 1
            product_stats.failed_items.append(f"{start_cell} {field_name}: {exc}")
            raise

    def read_cell(self, cell: str) -> str:
        self.goto_cell(cell)
        assert self.page is not None
        self.page.keyboard.press("Control+C")
        self.page.wait_for_timeout(150)
        return self.get_clipboard_text()

    def read_vertical_range(self, range_addr: str, expected_count: int, retries: int = 3) -> RangeReadResult:
        assert self.page is not None
        last_values: list[str] = []

        for attempt in range(1, retries + 1):
            self.clear_clipboard()
            self.goto_cell(range_addr)
            self.page.wait_for_timeout(500)
            self.page.keyboard.press("Control+C")

            text = ""
            values: list[str] = []
            changed = False
            for _ in range(20):
                self.page.wait_for_timeout(150)
                text = self.get_clipboard_text()
                changed = text != ""
                values = self._parse_vertical_clipboard(text)
                if changed and len(values) == expected_count:
                    break

            last_values = values
            if changed and len(values) == expected_count:
                return RangeReadResult(values=values, success=True)

        return RangeReadResult(values=last_values, success=False)

    @staticmethod
    def _parse_vertical_clipboard(text: str) -> list[str]:
        if text == "":
            return []
        return [line.split("\t")[0] if line else "" for line in text.splitlines()]

    @staticmethod
    def _target_range(start_cell: str, row_count: int) -> str:
        match = re.fullmatch(r"([A-Z]+)(\d+)", start_cell.strip().upper())
        if not match:
            return start_cell
        col, row_text = match.groups()
        row = int(row_text)
        return f"{col}{row}:{col}{row + row_count - 1}"

    def set_clipboard_text(self, text: str) -> None:
        if pyperclip is not None:
            pyperclip.copy(text)
            return
        assert self.page is not None
        self.context.grant_permissions(["clipboard-read", "clipboard-write"], origin="https://www.kdocs.cn")
        self.page.evaluate(
            """async value => {
                await navigator.clipboard.writeText(value);
            }""",
            text,
        )

    def clear_clipboard(self) -> None:
        self.set_clipboard_text("")

    def get_clipboard_text(self) -> str:
        if pyperclip is not None:
            return pyperclip.paste() or ""
        assert self.page is not None
        self.context.grant_permissions(["clipboard-read", "clipboard-write"], origin="https://www.kdocs.cn")
        value = self.page.evaluate(
            """async () => {
                return await navigator.clipboard.readText();
            }"""
        )
        return value or ""

    @staticmethod
    def _to_clipboard_text(value: Any) -> str:
        return "" if value is None else str(value)


class ChangeLogger:
    """写出人工可读修改清单。"""

    def save(self, path: Path, report: LoadedReport, plan: list[ChangeItem], safety_issues: list[str], dry_run: bool) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        lines: list[str] = []
        lines.append("在线表格修改清单")
        lines.append(f"DRY_RUN={dry_run}")
        lines.append(f"目标文档={DOC_URL}")
        lines.append(f"目标工作表关键字={TARGET_SHEET_KEYWORD}")
        lines.append(f"识别日期={report.dates.slash} / {report.dates.compact} / {report.dates.short}")
        lines.append(f"结构变化处理模式={STRUCTURAL_CHANGES_MODE}")
        lines.append("")
        lines.append("安全校验")
        if safety_issues:
            lines.append("状态=阻止写入")
            for issue in safety_issues:
                lines.append(f"- {issue}")
        else:
            lines.append("状态=通过")
        lines.append("")
        lines.append("结构调整清单")
        if report.structural_changes:
            lines.append("默认处理方式=人工先处理插行、剪切和重排序号；处理后再次运行 dry-run。")
            for idx, change in enumerate(report.structural_changes, start=1):
                lines.append(f"{idx}. 组合产品={change.product}")
                lines.append(f"   子基金名称={change.fund_name}")
                lines.append(f"   当前序号={change.current_index}")
                lines.append(f"   状态={change.status}")
                lines.append(f"   建议动作={change.action}")
                lines.append(f"   建议插入位置或移出位置={change.suggested_position}")
                lines.append(f"   调整前行数={change.before_count}")
                lines.append(f"   调整后预计行数={change.after_count}")
        else:
            lines.append("无未处理的新增/全赎/剪切/未持有结构变化。")
        lines.append("")
        lines.append("需要人工提供的固定单元格映射")
        lines.extend(self._missing_mapping_lines())
        lines.append("")
        lines.append("修改计划")
        for idx, item in enumerate(plan, start=1):
            lines.append(f"{idx}. 组合产品={item.product}")
            lines.append(f"   目标范围={item.target_range}")
            lines.append(f"   数据条数={item.row_count}")
            lines.append(f"   字段={', '.join(item.fields)}")
            lines.append(f"   状态={item.status}")
            lines.append(f"   是否可安全写入={'是' if item.safe_to_write else '否'}")
            lines.append(f"   原因={item.reason}")
            lines.append("   准备写入的数据=")
            lines.extend(self._format_data(item.data))
        path.write_text("\n".join(lines), encoding="utf-8")

    @staticmethod
    def append_write_stats(path: Path, stats: WriteStats) -> None:
        lines: list[str] = []
        lines.append("")
        lines.append("写入执行统计")
        lines.append(f"总检查单元格数={stats.checked_cells}")
        lines.append(f"相同并跳过数={stats.skipped_cells}")
        lines.append(f"实际写入单元格数={stats.written_cells}")
        lines.append(f"实际写入操作数={stats.write_operations}")
        lines.append(f"失败数={stats.failed}")
        lines.append("")
        lines.append("每个组合产品的跳过/写入情况")
        for product in ["表头"] + PRODUCT_ORDER:
            product_stats = stats.products.get(product)
            if not product_stats:
                continue
            lines.append(f"- {product}")
            lines.append(f"  检查={product_stats.checked_cells}")
            lines.append(f"  跳过={product_stats.skipped_cells}")
            lines.append(f"  写入单元格={product_stats.written_cells}")
            lines.append(f"  写入操作={product_stats.write_operations}")
            lines.append(f"  失败={product_stats.failed}")
            if product_stats.skipped_items:
                lines.append(f"  跳过项={'; '.join(product_stats.skipped_items)}")
            if product_stats.written_items:
                lines.append(f"  写入项={'; '.join(product_stats.written_items)}")
            if product_stats.diff_items:
                lines.append(f"  差异项={'; '.join(product_stats.diff_items)}")
            if product_stats.failed_items:
                lines.append(f"  失败项={'; '.join(product_stats.failed_items)}")
        with path.open("a", encoding="utf-8") as fh:
            fh.write("\n".join(lines))

    @staticmethod
    def _missing_mapping_lines() -> list[str]:
        missing: list[str] = []
        for field_name, cell in HEADER_CELLS.items():
            if not cell or (isinstance(cell, list) and not cell):
                missing.append(f"- HEADER_CELLS['{field_name}']")
        for product, field_map in SUMMARY_CELL_MAP.items():
            for field_name, cell in field_map.items():
                if not cell:
                    missing.append(f"- SUMMARY_CELL_MAP['{product}']['{field_name}']")
        for product, cell in MARKET_VALUE_START_CELLS.items():
            if not cell:
                missing.append(f"- MARKET_VALUE_START_CELLS['{product}']")
        for product, confirmed in KDOCS_SEQUENCE_CONFIRMED.items():
            if not confirmed:
                missing.append(f"- KDOCS_SEQUENCE_CONFIRMED['{product}'] 需要人工确认后改为 True")
        for product, expected in EXPECTED_MARKET_VALUE_ROWS.items():
            if expected is None:
                missing.append(f"- EXPECTED_MARKET_VALUE_ROWS['{product}'] 需要填写人工确认后的金山现有行数")
        return missing or ["- 无"]

    @staticmethod
    def _format_data(data: Any) -> list[str]:
        if isinstance(data, dict):
            return [f"     {key}: {value or []}" for key, value in data.items()]
        if isinstance(data, list):
            if not data:
                return ["     []"]
            return [f"     {value}" for value in data]
        return [f"     {data}"]


def build_plan(report: LoadedReport) -> tuple[list[ChangeItem], list[ChangeItem]]:
    auto_plan: list[ChangeItem] = []
    auto_plan.extend(HeaderUpdater().build_plan(report.dates))
    auto_plan.extend(SummaryBlockUpdater().build_plan(report))
    auto_plan.extend(MarketValueBlockUpdater().build_plan(report))
    manual_plan = ComplexFieldPlanner().build_plan(report)
    return auto_plan, manual_plan


def apply_auto_plan(auto_plan: list[ChangeItem]) -> WriteStats:
    header_items = [item for item in auto_plan if item.product == "表头"]
    summary_items = [item for item in auto_plan if item.product in PRODUCT_ORDER and item.fields != ["市值"]]
    market_items = [item for item in auto_plan if item.product in PRODUCT_ORDER and item.fields == ["市值"]]

    with KDocsBrowser(DOC_URL) as browser:
        browser.select_latest_sheet()
        HeaderUpdater().apply(browser, header_items)
        SummaryBlockUpdater().apply(browser, summary_items)
        MarketValueBlockUpdater().apply(browser, market_items)
        return browser.stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按固定顺序整块粘贴金山月报数据。默认只生成 DRY_RUN 修改清单。")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="只生成 output/playwright_修改清单.txt，不修改在线表格。默认行为。")
    mode.add_argument("--write", action="store_true", help="通过安全校验后真实写入在线表格。")
    parser.add_argument("--open-browser", action="store_true", help="DRY_RUN 时也打开在线表格并尝试自动切换目标工作表。")
    parser.add_argument("--yes", action="store_true", help="真实写入时跳过二次输入确认；仍必须通过安全校验。")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    dry_run = not args.write

    loader = MonthlyReportLoader(OUTPUT_FILE)
    report = loader.load()
    auto_plan, manual_plan = build_plan(report)
    full_plan = auto_plan + manual_plan
    safety_issues = SafetyValidator().validate(report, auto_plan)
    ChangeLogger().save(LOG_FILE, report, full_plan, safety_issues, dry_run=dry_run)

    print(f"已生成修改清单：{LOG_FILE}")
    print(f"DRY_RUN={dry_run}")
    print(f"识别日期：{report.dates.slash} / {report.dates.compact} / {report.dates.short}")
    if safety_issues:
        print("安全校验未通过，真实写入会被阻止：")
        for issue in safety_issues[:20]:
            print(f"- {issue}")
        if len(safety_issues) > 20:
            print(f"- 另有 {len(safety_issues) - 20} 条，详见修改清单")

    if dry_run:
        if args.open_browser:
            with KDocsBrowser(DOC_URL) as browser:
                browser.select_latest_sheet()
                input("DRY_RUN 已打开文档但不会写入。按回车关闭浏览器：")
        return 0

    if not DOC_URL:
        print("未配置 KDOCS_DOC_URL，已阻止真实写入。")
        return 2

    if safety_issues:
        print("存在风险或配置缺失，已阻止写入。")
        return 2

    if not args.yes:
        confirm = input("即将真实写入在线表格。输入 YES 才继续：")
        if confirm != "YES":
            print("已取消写入。")
            return 1

    write_stats = apply_auto_plan(auto_plan)
    ChangeLogger.append_write_stats(LOG_FILE, write_stats)
    print(f"总检查单元格数={write_stats.checked_cells}")
    print(f"相同并跳过数={write_stats.skipped_cells}")
    print(f"实际写入单元格数={write_stats.written_cells}")
    print(f"实际写入操作数={write_stats.write_operations}")
    print(f"失败数={write_stats.failed}")
    print("写入完成，请在在线表格中复核。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
