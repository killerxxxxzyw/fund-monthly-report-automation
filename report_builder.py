# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import PORTFOLIO_FUND_ORDER, PORTFOLIO_ORDER


def _write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]):
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["提示"])
        ws.append(["暂无数据"])
        return ws
    headers: List[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 48))
        ws.column_dimensions[get_column_letter(col)].width = max_len
    return ws


def _txn_summary_map(txn_summary: List[Dict[str, Any]]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    return {(row.get("组合产品", ""), row.get("底层产品名称", "")): row for row in txn_summary}


def build_nested_fill_rows(underlying_rows, txn_summary):
    txn_map = _txn_summary_map(txn_summary)
    seen = set()
    rows = []
    for underlying in underlying_rows:
        portfolio = underlying.get("组合产品", "")
        fund = underlying.get("底层产品名称", "")
        key = (portfolio, fund)
        seen.add(key)
        txn = txn_map.get(key, {})
        rows.append({
            "状态": "正常持有",
            "建议动作": "更新原有行",
            "组合产品": portfolio,
            "底层产品名称": fund,
            "底层产品市值": underlying.get("底层产品市值", 0),
            "申购金额": txn.get("申购金额", 0),
            "赎回金额": txn.get("赎回金额", 0),
            "分红金额": txn.get("分红金额", 0),
            "净申购": txn.get("净申购", 0),
            "已投金额调整=申购-赎回": txn.get("已投金额调整", 0),
            "申购文字": txn.get("申购文字_追加", ""),
            "赎回文字": txn.get("赎回文字_追加", ""),
            "分红文字": txn.get("分红文字_追加", ""),
            "来源估值表": underlying.get("来源文件", ""),
        })

    for key, txn in sorted(txn_map.items()):
        if key in seen:
            continue
        purchase = txn.get("申购金额", 0) or 0
        redemption = txn.get("赎回金额", 0) or 0
        dividend = txn.get("分红金额", 0) or 0
        adjustment = purchase - redemption
        if adjustment > 0:
            status, action = "新增认购_估值表未持有", "新增一行后重新编号"
        elif redemption > 0 or purchase > 0:
            status, action = "疑似全赎_估值表未持有", "人工确认后移至归档区"
        elif dividend > 0:
            status, action = "仅分红但估值表未持有", "人工确认名称或历史状态"
        else:
            status, action = "交易有但估值表未持有", "人工确认"
        rows.append({
            "状态": status,
            "建议动作": action,
            "组合产品": key[0],
            "底层产品名称": key[1],
            "底层产品市值": 0,
            "申购金额": purchase,
            "赎回金额": redemption,
            "分红金额": dividend,
            "净申购": adjustment,
            "已投金额调整=申购-赎回": adjustment,
            "申购文字": txn.get("申购文字_追加", ""),
            "赎回文字": txn.get("赎回文字_追加", ""),
            "分红文字": txn.get("分红文字_追加", ""),
            "来源估值表": "",
        })
    return rows


def build_kdocs_paste_rows(nested_rows):
    nested_map = {(row.get("组合产品", ""), row.get("底层产品名称", "")): row for row in nested_rows}
    result = []
    used = set()
    for portfolio, ordered_funds in PORTFOLIO_FUND_ORDER.items():
        for seq, (manager, fund) in enumerate(ordered_funds, start=1):
            key = (portfolio, fund)
            source = nested_map.get(key)
            used.add(key)
            result.append({
                "组合产品": portfolio,
                "序号": seq,
                "管理人": manager,
                "子基金": fund,
                "市值": source.get("底层产品市值", "") if source else "",
                "申购金额": source.get("申购金额", 0) if source else 0,
                "赎回金额": source.get("赎回金额", 0) if source else 0,
                "分红金额": source.get("分红金额", 0) if source else 0,
                "申购文字": source.get("申购文字", "") if source else "",
                "赎回文字": source.get("赎回文字", "") if source else "",
                "分红文字": source.get("分红文字", "") if source else "",
                "已投金额调整=申购-赎回": source.get("已投金额调整=申购-赎回", 0) if source else 0,
                "状态": source.get("状态", "模板原有产品_本期估值表未持有") if source else "模板原有产品_本期估值表未持有",
                "建议动作": source.get("建议动作", "人工确认") if source else "人工确认",
                "是否可直接粘贴市值": "是" if source and "未持有" not in str(source.get("状态", "")) else "需确认",
            })
    for row in nested_rows:
        key = (row.get("组合产品", ""), row.get("底层产品名称", ""))
        if key in used:
            continue
        portfolio = key[0]
        result.append({
            "组合产品": portfolio,
            "序号": len(PORTFOLIO_FUND_ORDER.get(portfolio, [])) + 1,
            "管理人": "",
            "子基金": key[1],
            "市值": row.get("底层产品市值", ""),
            "申购金额": row.get("申购金额", 0),
            "赎回金额": row.get("赎回金额", 0),
            "分红金额": row.get("分红金额", 0),
            "申购文字": row.get("申购文字", ""),
            "赎回文字": row.get("赎回文字", ""),
            "分红文字": row.get("分红文字", ""),
            "已投金额调整=申购-赎回": row.get("已投金额调整=申购-赎回", 0),
            "状态": row.get("状态", "新增产品"),
            "建议动作": "新增待插行；插入后重新编号",
            "是否可直接粘贴市值": "否，先在在线表格插行",
        })
    order = {name: idx for idx, name in enumerate(PORTFOLIO_ORDER)}
    result.sort(key=lambda row: (order.get(row.get("组合产品", ""), 999), row.get("序号", 9999)))
    return result


def build_text_rows(txn_details):
    rows = []
    for item in sorted(
        [row for row in txn_details if not row.get("错误")],
        key=lambda row: (row.get("日期", ""), row.get("组合产品", ""), row.get("底层产品名称", "")),
    ):
        rows.append({
            "日期": item.get("日期", ""),
            "交易类型": item.get("交易类型", ""),
            "组合产品": item.get("组合产品", ""),
            "底层产品名称": item.get("底层产品名称", ""),
            "金额": item.get("金额", 0),
            "金额_万元": item.get("金额_万元", 0),
            "可复制文字": item.get("在线表格文字", ""),
        })
    return rows


def build_report(output_path: Path, valuation_summary, underlying_rows, txn_details, txn_summary, txn_summary_by_portfolio):
    wb = Workbook()
    wb.remove(wb.active)

    nested_rows = build_nested_fill_rows(underlying_rows, txn_summary)
    kdocs_rows = build_kdocs_paste_rows(nested_rows)

    _write_sheet(wb, "在线表格粘贴版", kdocs_rows)
    _write_sheet(wb, "填表用_底层产品", nested_rows)
    _write_sheet(wb, "估值汇总", valuation_summary)
    _write_sheet(wb, "底层产品市值", underlying_rows)
    _write_sheet(wb, "交易明细", txn_details)
    _write_sheet(wb, "交易汇总_底层产品", txn_summary)
    _write_sheet(wb, "交易汇总_组合产品", txn_summary_by_portfolio)
    _write_sheet(wb, "文字版_申赎分红", build_text_rows(txn_details))

    txn_map = {row.get("组合产品"): row for row in txn_summary_by_portfolio}
    valuation_map = {row.get("组合产品", ""): row for row in valuation_summary}
    summary_rows = []
    for portfolio in PORTFOLIO_ORDER:
        valuation = valuation_map.get(portfolio, {})
        txn = txn_map.get(portfolio, {})
        summary_rows.append({
            "组合产品": portfolio,
            "净值": valuation.get("基金单位净值", ""),
            "产品成本规模_实收资本": valuation.get("产品成本规模_实收资本", ""),
            "产品市值规模_基金资产净值": valuation.get("产品市值规模_基金资产净值", ""),
            "剩余资金_万元": valuation.get("剩余资金_万元", ""),
            "剩余资金备注": valuation.get("剩余资金备注", ""),
            "申购合计": txn.get("申购合计", 0),
            "赎回合计": txn.get("赎回合计", 0),
            "分红合计": txn.get("分红合计", 0),
            "净申购合计": txn.get("净申购合计", 0),
        })
    _write_sheet(wb, "填表用_组合汇总", summary_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
