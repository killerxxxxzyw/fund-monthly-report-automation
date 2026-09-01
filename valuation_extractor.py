# -*- coding: utf-8 -*-
"""Extract portfolio-level and underlying-fund values from valuation sheets."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from config import (
    FIELD_BANK_DEPOSIT,
    FIELD_NAV,
    FIELD_NET_ASSET,
    FIELD_OTC_MONEY_FUND,
    FIELD_PAID_IN_CAPITAL,
    UNDERLYING_CODE_PREFIX,
)
from utils import extract_portfolio_from_text, normalize_text, to_number


def _get_header_indexes(ws) -> Dict[str, int]:
    for row in ws.iter_rows():
        values = [normalize_text(cell.value) for cell in row]
        if "科目代码" in values and "科目名称" in values:
            return {value: idx for idx, value in enumerate(values, start=1) if value}
    return {}


def _find_row_by_exact_name(ws, keyword: str, name_col: int) -> int:
    target = normalize_text(keyword)
    for row in range(1, ws.max_row + 1):
        if normalize_text(ws.cell(row, name_col).value) == target:
            return row
    return 0


def _find_row_contains(ws, keyword: str) -> Tuple[int, int]:
    target = normalize_text(keyword)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if target and target in normalize_text(ws.cell(row, col).value):
                return row, col
    return 0, 0


def _extract_portfolio_name(ws, file_path: Path) -> str:
    for row in range(1, min(ws.max_row, 5) + 1):
        row_text = " ".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1))
        portfolio = extract_portfolio_from_text(row_text)
        if portfolio:
            return portfolio
    return extract_portfolio_from_text(file_path.name) or file_path.stem


def _value_from_row(ws, row: int, preferred_col: int = 0) -> float:
    if row <= 0:
        return 0.0
    if preferred_col:
        raw = ws.cell(row, preferred_col).value
        if raw not in (None, ""):
            return to_number(raw)
    numbers = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row, col).value
        if isinstance(raw, (int, float)):
            numbers.append(to_number(raw))
    return numbers[-1] if numbers else 0.0


def extract_valuation_file(file_path: Path) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = _get_header_indexes(ws)
    if not headers:
        raise ValueError("No valuation header row containing 科目代码 / 科目名称 was found")

    code_col = headers.get("科目代码", 1)
    name_col = headers.get("科目名称", 2)
    market_value_col = headers.get("市值", ws.max_column)
    portfolio = _extract_portfolio_name(ws, file_path)

    nav_row, nav_col = _find_row_contains(ws, FIELD_NAV)
    nav = 0.0
    if nav_row:
        row_numbers = []
        for col in range(1, ws.max_column + 1):
            if col == nav_col:
                continue
            raw = ws.cell(nav_row, col).value
            if isinstance(raw, (int, float)):
                row_numbers.append(to_number(raw))
        nav = row_numbers[-1] if row_numbers else 0.0

    paid_row = _find_row_by_exact_name(ws, FIELD_PAID_IN_CAPITAL, name_col)
    net_asset_row = _find_row_by_exact_name(ws, FIELD_NET_ASSET, name_col)
    bank_row = _find_row_by_exact_name(ws, FIELD_BANK_DEPOSIT, name_col)
    otc_row = _find_row_by_exact_name(ws, FIELD_OTC_MONEY_FUND, name_col)

    paid_in_capital = _value_from_row(ws, paid_row, market_value_col)
    net_asset = _value_from_row(ws, net_asset_row, market_value_col)
    bank_deposit = _value_from_row(ws, bank_row, market_value_col)
    otc_money_fund = _value_from_row(ws, otc_row, market_value_col)

    summary = {
        "组合产品": portfolio,
        "估值表文件": file_path.name,
        "基金单位净值": nav,
        "产品成本规模_实收资本": paid_in_capital,
        "产品市值规模_基金资产净值": net_asset,
        "银行存款": bank_deposit,
        "场外货币基金": otc_money_fund,
        "剩余资金_万元": (bank_deposit + otc_money_fund) / 10000,
        "剩余资金备注": f"含{otc_money_fund / 10000:.2f}万现金管理类基金" if otc_money_fund else "",
    }

    underlyings: List[Dict[str, Any]] = []
    for row in range(1, ws.max_row + 1):
        code = str(ws.cell(row, code_col).value or "").strip()
        if not code.startswith(UNDERLYING_CODE_PREFIX):
            continue
        fund_name = str(ws.cell(row, name_col).value or "").strip()
        if not fund_name or fund_name in {"成本", "估值增值"}:
            continue
        underlyings.append({
            "组合产品": portfolio,
            "底层产品名称": fund_name,
            "底层产品市值": to_number(ws.cell(row, market_value_col).value),
            "来源文件": file_path.name,
        })
    return summary, underlyings


def extract_all_valuations(files: List[Path]):
    summaries: List[Dict[str, Any]] = []
    underlyings: List[Dict[str, Any]] = []
    for file_path in files:
        try:
            summary, rows = extract_valuation_file(file_path)
            summaries.append(summary)
            underlyings.extend(rows)
        except Exception as exc:
            summaries.append({"组合产品": file_path.stem, "估值表文件": file_path.name, "错误": str(exc)})
    return summaries, underlyings
